import json
import os
import requests
from pathlib import Path
from typing import List, Generator, Dict, Any, Literal, Optional
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from tqdm.auto import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

from .exceptions import (
    UnihraError, UnihraApiError, UnihraConnectionError,
    UnihraValidationError, UnihraDependencyError, UnihraStorageError, raise_for_error_code
)

BASE_URL = "https://unihra.ru"

ACTION_MAP = {
    "Добавить": "add",
    "Увеличить": "increase",
    "Уменьшить": "decrease",
    "Ок": "ok",
    "Ничего не делать": "ok"
}

class UnihraClient:
    """
    Official Python Client for Unihra API with local storage support.
    """

    def __init__(self, api_key: str, base_url: str = BASE_URL, max_retries: int = 0, storage_dir: str = "unihra_results"):
        self.base_url = base_url.rstrip('/')
        self.api_v1 = f"{self.base_url}/api/v1"
        self.storage_path = Path(storage_dir)
        self.session = requests.Session()

        self.session.headers.update({
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "UnihraPythonSDK/1.7.0"
        })

        if max_retries > 0:
            retry_strategy = Retry(
                total=max_retries,
                backoff_factor=2,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["POST", "GET"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            self.session.mount("https://", adapter)
            self.session.mount("http://", adapter)

    def health(self) -> Dict[str, Any]:
        """Check API availability."""
        try:
            resp = self.session.get(f"{self.api_v1}/health")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Health check failed: {e}")

    def get_page_structure(self, task_id: str) -> List[Dict[str, Any]]:
        """
        Fetch detailed page structure list (Own Page + Competitors).
        """
        try:
            resp = self.session.get(f"{self.api_v1}/report/structure/{task_id}")
            resp.raise_for_status()
            data = resp.json()

            # API returns a List of objects. Normalize each one.
            if isinstance(data, list):
                return[self._normalize_keys(item) for item in data]
            return[]
        except requests.exceptions.RequestException:
            return[]

    def analyze(
        self,
        own_page: str,
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: Literal['ru', 'en'] = 'ru',
        url_cookies: Optional[Dict[str, str]] = None,
        triplet_analysis: bool = False,
        verbose: bool = False
    ) -> Dict[str, Any]:
        """
        Start a full analysis task and wait for completion.

        :param own_page: URL of your landing page.
        :param competitors: List of competitor URLs.
        :param queries: List of target keywords.
        :param lang: 'ru' or 'en'.
        :param url_cookies: Dictionary mapping URL -> Cookie String.
        :param triplet_analysis: Enable extended Knowledge Graph extraction (5 credits vs 1).
        :param verbose: If True, displays a progress bar.
        """
        last_event = {}
        pbar = None

        if verbose:
            if TQDM_AVAILABLE:
                desc = "Analyzing SEO (+ Triplets)" if triplet_analysis else "Analyzing SEO"
                pbar = tqdm(total=100, desc=desc, unit="%")
            else:
                print("Note: Install 'tqdm' to see a visual progress bar.")

        try:
            for event in self.analyze_stream(own_page, competitors, queries, lang, url_cookies, triplet_analysis):
                last_event = event

                if pbar:
                    state = event.get("state")
                    progress = event.get("progress", 0)

                    if isinstance(progress, (int, float)):
                        pbar.n = int(progress)
                        pbar.refresh()

                    if state in ["PROCESSING", "PROGRESS"]:
                        msg = "Processing"
                        details = event.get("details", {})
                        if isinstance(details, dict) and "message" in details:
                            msg = details["message"][:40]
                        pbar.set_description(f"{msg}")
                    elif state == "SUCCESS":
                        pbar.set_description("Completed ✅")
                        pbar.n = 100
                        pbar.refresh()

                if event.get("state") == "SUCCESS":
                    return event.get("result", {})

        except Exception as e:
            if pbar:
                pbar.set_description("Failed ❌")
                pbar.close()
            raise e
        finally:
            if pbar:
                pbar.close()

        return last_event

    def _strip_id_recursively(self, obj: Any) -> Any:
        """Рекурсивная очистка технических ID из данных перед сохранением."""
        tech_keys = {"analysis_id", "task_id", "block_id", "id"}
        if isinstance(obj, dict):
            return {k: self._strip_id_recursively(v) for k, v in obj.items() if k not in tech_keys}
        if isinstance(obj, list):
            return [self._strip_id_recursively(i) for i in obj]
        return obj

    def analyze_and_save(self, **kwargs) -> Dict[str, Any]:
        """
        Runs analysis, cleans IDs and splits result into local JSON files.
        Returns a manifest with file paths and analysis ID.
        """
        result = self.analyze(**kwargs)
        analysis_id = result.get("_meta", {}).get("task_id", "unknown")

        target_dir = self.storage_path / analysis_id
        target_dir.mkdir(parents=True, exist_ok=True)

        manifest = {
            "analysis_id": analysis_id,
            "files": {}
        }

        # Mapping internal keys to segment filenames
        segments = {
            "gaps": "umbrella_analysis",
            "words": "block_comparison",
            "ngrams": "ngrams_analysis",
            "anchors": "anchors_analysis",
            "triplets": "triplets_analysis",
            "structure": "page_structure"
        }

        try:
            for file_key, data_key in segments.items():
                if data_key == "umbrella_analysis":
                    data = result.get("umbrella_analysis") or result.get("semantic_context_analysis") or []
                elif data_key == "triplets_analysis":
                    data = result.get(data_key, {})
                else:
                    data = result.get(data_key, [])

                # Применяем очистку перед сохранением
                cleaned_data = self._strip_id_recursively(data)

                file_path = target_dir / f"{file_key}.json"
                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(cleaned_data, f, ensure_ascii=False, indent=2)
                manifest["files"][file_key] = str(file_path.absolute())

            return manifest
        except Exception as e:
            raise UnihraStorageError(f"Failed to save analysis segments: {e}")

    def analyze_stream(
        self,
        own_page: str,
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: str = 'ru',
        url_cookies: Optional[Dict[str, str]] = None,
        triplet_analysis: bool = False
    ) -> Generator[Dict, None, None]:
        if not competitors:
            raise UnihraValidationError("Competitor list cannot be empty.")

        payload = {
            "own_page": own_page,
            "competitor_urls": competitors,
            "queries": queries or[],
            "lang": lang,
            "url_cookies": url_cookies or {},
            "triplet_analysis": bool(triplet_analysis)
        }

        try:
            resp = self.session.post(f"{self.api_v1}/process", json=payload)

            if resp.status_code == 401:
                raise UnihraApiError("Invalid API Key or unauthorized access", code=401)
            resp.raise_for_status()

            task_id = resp.json().get("task_id")
            if not task_id:
                raise UnihraApiError("API response missing 'task_id'")

            stream_url = f"{self.api_v1}/process/status/{task_id}"

            with self.session.get(stream_url, stream=True) as s_resp:
                s_resp.raise_for_status()

                for line in s_resp.iter_lines():
                    if not line:
                        continue

                    if line.startswith(b'data: '):
                        try:
                            decoded_line = line[6:].decode('utf-8')
                            data = json.loads(decoded_line)
                            state = data.get("state")

                            if state == "FAILURE":
                                error_obj = data.get("error")
                                code = error_obj.get("code", 9999) if isinstance(error_obj, dict) else 9999
                                msg = error_obj.get("message", "Unknown error") if isinstance(error_obj, dict) else "Unknown error"
                                raise_for_error_code(code, msg, data)

                            if state == "SUCCESS":
                                raw_result = data.get("result", {})
                                normalized_result = self._normalize_keys(raw_result)
                                normalized_result["_meta"] = {
                                    "task_id": task_id,
                                    "triplet_analysis": bool(triplet_analysis),
                                    "credits_spent": 5 if triplet_analysis else 1,
                                }

                                if lang == 'en':
                                    final_result = self._translate_action_values(normalized_result)
                                else:
                                    final_result = normalized_result

                                # Automatically fetch and attach page structure
                                structure_data = self.get_page_structure(task_id)
                                if structure_data:
                                    final_result['page_structure'] = structure_data

                                data["result"] = final_result
                                yield data
                                break

                            yield data

                        except json.JSONDecodeError:
                            continue

        except requests.exceptions.RetryError:
            raise UnihraConnectionError("Max retries exceeded. The service might be temporarily unavailable.")
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Network error: {e}")

    def _normalize_keys(self, data: Any) -> Any:
        if isinstance(data, dict):
            return {
                k.lower().replace(" ", "_").replace("-", "_"): self._normalize_keys(v)
                for k, v in data.items()
            }
        elif isinstance(data, list):
            return [self._normalize_keys(i) for i in data]
        return data

    def _translate_action_values(self, result: Dict[str, Any]) -> Dict[str, Any]:
        if "block_comparison" in result and isinstance(result["block_comparison"], list):
            for item in result["block_comparison"]:
                if "action_needed" in item:
                    russian_action = item["action_needed"]
                    item["action_needed"] = ACTION_MAP.get(russian_action, russian_action)
        return result

    def _flatten_structure_list(self, structure_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten list of structures for DataFrame (Table view)."""
        flat_rows =[]
        for item in structure_list:
            flat_item = {'url': item.get('url')}
            for section in['metrics', 'content', 'meta_tags']:
                if section in item:
                    for k, v in item[section].items():
                        flat_item[k] = v
            flat_rows.append(flat_item)
        return flat_rows

    def _flatten_triplets_entities(self, entities: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten knowledge-graph entities into one row per triplet for tabular view."""
        flat_rows = []
        for ent in entities or []:
            subject = ent.get("subject")
            tier = ent.get("tier")
            sources_count = ent.get("sources_count")
            triplets_count = ent.get("triplets_count")
            for tr in ent.get("triplets") or []:
                flat_rows.append({
                    "subject":         subject,
                    "tier":            tier,
                    "predicate":       tr.get("predicate"),
                    "object":          tr.get("object"),
                    "sources":         ", ".join(tr.get("sources") or []),
                    "sources_count":   sources_count,
                    "triplets_count":  triplets_count,
                })
        return flat_rows

    def _flatten_triplets_gaps(self, gaps: Dict[str, List[Any]]) -> List[Dict[str, Any]]:
        """Flatten gap subjects (critical / important / unique) into a single table."""
        flat_rows = []
        for severity in ("critical", "important", "unique"):
            for item in (gaps or {}).get(severity) or []:
                if isinstance(item, dict):
                    row = {"severity": severity}
                    row.update(item)
                else:
                    row = {"severity": severity, "subject": item}
                flat_rows.append(row)
        return flat_rows

    def get_dataframe(self, result: Dict[str, Any], section: str = "block_comparison"):
        """
        Convert a specific section of the result into a Pandas DataFrame.
        Supported sections: block_comparison, ngrams_analysis, umbrella_analysis,
        anchors_analysis, page_structure, triplets_analysis, triplets_gaps.
        Note: umbrella_analysis was previously named semantic_context_analysis — both are accepted.
        """
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is not installed. Run: pip install unihra[full]")

        normalized_section = section.lower().replace(" ", "_").replace("-", "_")

        if normalized_section == "page_structure":
            data = result.get("page_structure",[])
            if not data:
                return pd.DataFrame()
            flat_list = self._flatten_structure_list(data)
            return pd.DataFrame(flat_list)

        if normalized_section == "triplets_analysis":
            entities = (result.get("triplets_analysis") or {}).get("entities") or []
            flat_list = self._flatten_triplets_entities(entities)
            return pd.DataFrame(flat_list)

        if normalized_section in ("triplets_gaps", "umbrella_analysis", "semantic_context_analysis"):
            if normalized_section == "triplets_gaps":
                ta = result.get("triplets_analysis") or {}
                gaps = ta.get("missing_triplets") or ta.get("gaps") or {}
                flat_list = self._flatten_triplets_gaps(gaps)
                return pd.DataFrame(flat_list)
            # umbrella_analysis / semantic_context_analysis — fall through to generic path below
            data = result.get("umbrella_analysis") or result.get("semantic_context_analysis") or result.get("semantic_context_gaps") or []
            return pd.DataFrame(data)

        data = result.get(normalized_section,[])
        return pd.DataFrame(data)

    def save_report(self, result: Dict[str, Any], filename: str = "report.xlsx", style_output: bool = True):
        """
        Saves the analysis result to Excel or CSV.
        """
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is required. Run: pip install unihra[report]")

        df_blocks = pd.DataFrame(result.get("block_comparison", []))
        df_ngrams = pd.DataFrame(result.get("ngrams_analysis") or result.get("n_grams_analysis") or [])
        df_gaps = pd.DataFrame(result.get("umbrella_analysis") or result.get("semantic_context_analysis") or result.get("semantic_context_gaps") or [])
        df_anchors = pd.DataFrame(result.get("anchors_analysis", []))
        triplets_data = result.get("triplets_analysis") or {}
        structure_data = result.get("page_structure",[])

        if filename.endswith(".csv"):
            if not df_blocks.empty:
                df_blocks = self._reorder_tech_columns(df_blocks)
            df_blocks.to_csv(filename, index=False, encoding='utf-8-sig')
        else:
            try:
                import openpyxl
            except ImportError:
                raise UnihraDependencyError("Library 'openpyxl' is required for Excel export.")

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 0. Anchors Analysis
                if not df_anchors.empty:
                    sheet = "Anchors"
                    df_anchors_ordered = self._reorder_tech_columns(df_anchors)
                    df_anchors_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_anchors_ordered, sheet_type="anchors")

                # 1. Page Structure
                if structure_data:
                    sheet = "Page Structure"
                    flat_struct = self._flatten_structure_list(structure_data)
                    df_struct = pd.DataFrame(flat_struct)

                    cols = df_struct.columns.tolist()
                    if 'url' in cols:
                        cols.insert(0, cols.pop(cols.index('url')))
                        df_struct = df_struct[cols]

                    df_struct.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_struct, sheet_type="structure")

                # 2. Semantic Gaps
                if not df_gaps.empty:
                    sheet = "Semantic Gaps"
                    desired_cols =['lemma', 'recommendation', 'context_snippet', 'gap', 'coverage_percent', 'competitor_avg_score', 'own_score']
                    existing_cols =[c for c in desired_cols if c in df_gaps.columns]
                    other_cols =[c for c in df_gaps.columns if c not in desired_cols]

                    df_gaps_ordered = df_gaps[existing_cols + other_cols]
                    df_gaps_ordered = self._reorder_tech_columns(df_gaps_ordered)

                    df_gaps_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_gaps_ordered, sheet_type="gaps")

                # 3. Word Analysis
                if not df_blocks.empty:
                    sheet = "Word Analysis"
                    df_blocks_ordered = self._reorder_tech_columns(df_blocks)
                    df_blocks_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_blocks_ordered, sheet_type="word_analysis")

                # 4. N-Grams
                if not df_ngrams.empty:
                    sheet = "N-Grams"
                    df_ngrams_ordered = self._reorder_tech_columns(df_ngrams)
                    df_ngrams_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_ngrams_ordered, sheet_type="ngrams")

                # 5. Triplets — Knowledge Graph (only if extended analysis was requested)
                if triplets_data and isinstance(triplets_data, dict):
                    entities = triplets_data.get("entities") or []
                    gaps_block = triplets_data.get("missing_triplets") or triplets_data.get("gaps") or {}

                    if entities:
                        sheet = "Triplets"
                        df_tr = pd.DataFrame(self._flatten_triplets_entities(entities))
                        df_tr_ordered = self._reorder_tech_columns(df_tr)
                        df_tr_ordered.to_excel(writer, sheet_name=sheet, index=False)
                        if style_output: self._style_worksheet(writer.sheets[sheet], df_tr_ordered, sheet_type="triplets")

                    if gaps_block:
                        sheet = "Triplets Gaps"
                        df_tg = pd.DataFrame(self._flatten_triplets_gaps(gaps_block))
                        if not df_tg.empty:
                            df_tg_ordered = self._reorder_tech_columns(df_tg)
                            df_tg_ordered.to_excel(writer, sheet_name=sheet, index=False)
                            if style_output: self._style_worksheet(writer.sheets[sheet], df_tg_ordered, sheet_type="triplets_gaps")

    def get_limits(self) -> Dict[str, Any]:
        """GET /api/v1/key/limits — current API key usage limits and remaining balance."""
        try:
            resp = self.session.get(f"{self.api_v1}/key/limits")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"get_limits failed: {e}")

    def list_analyses(self) -> List[Dict[str, Any]]:
        """GET /api/v1/analyses — list analyses saved under this API key."""
        try:
            resp = self.session.get(f"{self.api_v1}/analyses")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"list_analyses failed: {e}")

    def get_analysis(self, task_id: str) -> Dict[str, Any]:
        """GET /api/v1/analyses/{task_id} — fetch a saved analysis result by task ID."""
        try:
            resp = self.session.get(f"{self.api_v1}/analyses/{task_id}")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"get_analysis failed: {e}")

    def share_analysis(self, task_id: str) -> Dict[str, Any]:
        """POST /api/v1/analyses/{task_id}/share — create a public share link for an analysis."""
        try:
            resp = self.session.post(f"{self.api_v1}/analyses/{task_id}/share")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"share_analysis failed: {e}")

    def unshare_analysis(self, task_id: str) -> None:
        """DELETE /api/v1/analyses/{task_id}/share — revoke the share link for an analysis."""
        try:
            resp = self.session.delete(f"{self.api_v1}/analyses/{task_id}/share")
            resp.raise_for_status()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"unshare_analysis failed: {e}")

    def _reorder_tech_columns(self, df):
        try:
            import pandas as pd
            if not isinstance(df, pd.DataFrame) or df.empty: return df
            tech_cols = ['id', 'block_id', 'analysis_id']
            existing_tech = [c for c in df.columns if c in tech_cols]
            main_cols =[c for c in df.columns if c not in tech_cols]
            return df[main_cols + existing_tech] if existing_tech else df
        except ImportError:
            return df

    def _style_worksheet(self, worksheet, df, sheet_type="generic"):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="363636", end_color="363636", fill_type="solid")
        tech_header_font = Font(bold=True, color="000000")

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        tech_cols = ['id', 'block_id', 'analysis_id']

        # Format Headers
        for cell in worksheet[1]:
            col_name = str(cell.value) if cell.value else ""
            if col_name in tech_cols:
                cell.font = tech_header_font
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width
        for idx, col in enumerate(df.columns):
            if col in tech_cols:
                worksheet.column_dimensions[get_column_letter(idx + 1)].hidden = True
                continue
            max_len = max([len(str(s)) for s in df[col].astype(str).values] + [len(col)])
            final_width = min(max_len + 2, 70)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = final_width

        col_map = {name: i + 1 for i, name in enumerate(df.columns)}

        if sheet_type == "structure":
            for col_name in['url', 'h1_heading', 'meta_title', 'meta_description', 'heading_structure_raw']:
                if col_name in col_map:
                    idx = col_map[col_name]
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=idx).alignment = Alignment(wrap_text=True)

        elif sheet_type == "anchors":
            if 'frequency_own' in col_map and 'anchor' in col_map:
                f_idx = col_map['frequency_own']
                a_idx = col_map['anchor']
                for row in range(2, worksheet.max_row + 1):
                    val = worksheet.cell(row=row, column=f_idx).value
                    try:
                        is_missing = float(val) == 0 if val is not None else True
                    except (ValueError, TypeError):
                        is_missing = True
                    worksheet.cell(row=row, column=a_idx).fill = red_fill if is_missing else green_fill

            # Set wrap text for links column (can contain multiple URLs)
            if 'links' in col_map:
                idx = col_map['links']
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=idx).alignment = Alignment(wrap_text=True)

        elif sheet_type == "gaps":
            if 'own_score' in col_map and 'lemma' in col_map:
                score_idx = col_map['own_score']
                lemma_idx = col_map['lemma']
                for row in range(2, worksheet.max_row + 1):
                    score_val = worksheet.cell(row=row, column=score_idx).value
                    try:
                        is_missing = float(score_val) == 0 if score_val is not None else True
                    except (ValueError, TypeError):
                        is_missing = True
                    worksheet.cell(row=row, column=lemma_idx).fill = red_fill if is_missing else green_fill

        elif sheet_type == "triplets":
            # Highlight subject column by tier importance
            tier_color = {
                "core":       red_fill,        # most important — highlight strongly
                "main":       amber_fill,
                "additional": green_fill,
                "unique":     PatternFill(fill_type=None),
            }
            if 'tier' in col_map and 'subject' in col_map:
                tier_idx = col_map['tier']
                subj_idx = col_map['subject']
                for row in range(2, worksheet.max_row + 1):
                    tier_val = worksheet.cell(row=row, column=tier_idx).value
                    fill = tier_color.get(str(tier_val).lower()) if tier_val else None
                    if fill:
                        worksheet.cell(row=row, column=subj_idx).fill = fill
            # Wrap long object / sources cells
            for col_name in ('object', 'sources', 'predicate'):
                if col_name in col_map:
                    idx = col_map[col_name]
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=idx).alignment = Alignment(wrap_text=True)

        elif sheet_type == "triplets_gaps":
            # Colour-code severity column (critical = red, important = amber, unique = green)
            severity_color = {
                "critical":  red_fill,
                "important": amber_fill,
                "unique":    green_fill,
            }
            if 'severity' in col_map:
                sev_idx = col_map['severity']
                for row in range(2, worksheet.max_row + 1):
                    sev_val = worksheet.cell(row=row, column=sev_idx).value
                    fill = severity_color.get(str(sev_val).lower()) if sev_val else None
                    if fill:
                        worksheet.cell(row=row, column=sev_idx).fill = fill

        else:
            target_cols =[]
            if sheet_type == "word_analysis":
                target_names = ["word", "lemma"]
                target_cols = [col_map[c] for c in target_names if c in col_map]
            elif sheet_type == "ngrams":
                target_names = ["ngram"]
                target_cols = [col_map[c] for c in target_names if c in col_map]

            bool_col = 'present_on_own_page'
            if bool_col not in col_map and 'present_in_own' in col_map:
                 bool_col = 'present_in_own'

            if bool_col in col_map and target_cols:
                bool_idx = col_map[bool_col]
                for row in range(2, worksheet.max_row + 1):
                    is_present = worksheet.cell(row=row, column=bool_idx).value
                    fill_color = green_fill if is_present is True else red_fill if is_present is False else None
                    if fill_color:
                        for t_idx in target_cols:
                            worksheet.cell(row=row, column=t_idx).fill = fill_color
